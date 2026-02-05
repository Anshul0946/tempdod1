import re
import json
import openpyxl
from typing import Dict, Any, List, Optional
from .config import ProcessingContext
from .api_manager import APIManager

class Mapper:
    def __init__(self, context: ProcessingContext, api_manager: APIManager):
        self.context = context
        self.api = api_manager

    def _normalize_name(self, s: str) -> str:
        return re.sub(r"[^0-9a-zA-Z]", "", s).lower()

    def resolve_expression(self, expr: str) -> Any:
        # Simple resolution logic similar to original, but using context dict
        data_map = self.context.to_dict()
        
        expr = expr.strip()
        m = re.match(r"^([A-Za-z_]\w*)(.*)$", expr)
        if not m:
            return None
        
        base_raw = m.group(1)
        rest = m.group(2) or ""

        # Normalize base key lookup
        base_key = None
        norm_map = {self._normalize_name(k): k for k in data_map.keys()}
        base_key = norm_map.get(self._normalize_name(base_raw))
        
        if not base_key:
            return None

        current_obj = data_map[base_key]
        
        # Parse brackets: ['key'] or ["key"]
        keys = re.findall(r"\[['\"]([^'\"]+)['\"]\]", rest)
        
        for k in keys:
            if isinstance(current_obj, dict):
                # Case insensitive lookup
                found = None
                for real_k in current_obj.keys():
                    if self._normalize_name(real_k) == self._normalize_name(k):
                        found = real_k
                        break
                if found:
                    current_obj = current_obj[found]
                else:
                    return None
            else:
                 return None
        
        return current_obj

    def map_to_excel(self, xlsx_path: str):
        self.context.log("[MAPPER] Scanning workbook for BOLD+RED expressions...")
        try:
            wb = openpyxl.load_workbook(xlsx_path)
            sheet = wb.active
            
            cells_to_update = []
            
            # Limit scan range for performance
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=16):
                for cell in row:
                    val = cell.value
                    if not val or not isinstance(val, str):
                        continue
                    
                    # Check for Red Bold
                    font = cell.font
                    if not font or not font.bold:
                        continue
                    col = font.color
                    if not col or not col.rgb or str(col.rgb).upper()[-6:] != "FF0000":
                        continue

                    # Clean expression
                    expr = val.strip()
                    if (expr.startswith('"') and expr.endswith('"')) or (expr.startswith("'") and expr.endswith("'")):
                        expr = expr[1:-1].strip()
                        
                    self.context.extract_text.append(expr)
                    
                    # Resolve
                    resolved_val = self.resolve_expression(expr)
                    
                    if resolved_val is not None:
                         cells_to_update.append((cell, resolved_val))
                    else:
                        # Rule 3 Trigger: strict AI re-check would go here (simplified for now to NULL)
                        cell.value = "NULL" 

            # Apply updates
            for cell, val in cells_to_update:
                if isinstance(val, (dict, list)):
                    cell.value = json.dumps(val)
                else:
                    cell.value = val

            wb.save(xlsx_path)
            self.context.log(f"[MAPPER] Workbook updated: {xlsx_path}")
            
        except Exception as e:
            self.context.log(f"[ERROR] Mapping failed: {e}")
